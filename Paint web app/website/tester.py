from openpyxl import Workbook
import openpyxl
import sqlite3

wb = Workbook()

path = "C:\\Users\\Gustaf\\Downloads\\paints.xlsx"

wb = openpyxl.load_workbook(path, data_only=True)

ws = wb.active



conn = sqlite3.connect('database4.db')
c = conn.cursor()

for row in ws.iter_rows(min_row=1, max_col=3, values_only=True):
    

    c.execute('''
            INSERT INTO paints(paint_name, color, type) 
            VALUES 
                (?, ?, ?)''', (row[0], row[1], row[2])
            )

         


conn.commit()
conn.close()



